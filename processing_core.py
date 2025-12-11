# -*- coding: utf-8 -*-
"""
processing_core.py - GY Batch System 핵심 처리 모듈 (최종 수정 버전)

S1: 엑셀 데이터 기반 작업 사진 리네임
S2: 엑셀 데이터 기반 내품 사진 리네임
"""

import os
import re
import shutil
import unicodedata
from pathlib import Path
from openpyxl import load_workbook
from typing import List, Tuple, Dict, Any 

# ⭐ 오류 수정: 상대 경로 임포트 대신 절대 경로 사용 ⭐
# processing_core.py가 main.py와 같은 레벨에 있는 config.py를 찾도록 수정
try:
    from config import get_config
    OUTPUT_POLICY = get_config().get("OUTPUT_POLICY", "new_folder") 
except ImportError:
    OUTPUT_POLICY = "new_folder" 
    

# -------------------------------------------------------------
# 1) 공통 유틸리티
# -------------------------------------------------------------

IMG_WHITELIST = ('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tif', '.tiff')
FORBIDDEN_CHARS = r'<>:"/\\|?*'

def _nz(x, default=""):
    """Null/None/NaN 값 처리 및 빈 문자열 반환"""
    x = "" if x is None else str(x)
    x = x.strip()
    return default if x.lower() == "nan" or not x else x

def sanitize(name: str) -> str:
    """파일명으로 사용 불가능한 문자열을 정제하고 길이 제한을 둠"""
    s = unicodedata.normalize('NFC', str(name)).strip()
    s = s.replace(" ", "")
    s = re.sub(f"[{re.escape(FORBIDDEN_CHARS)}]", "_", s)
    s = re.sub(r"\s+", " ", s).strip(" .")
    # 파일명 최대 길이 제한 (180자)
    return s[:180]

def ensure_unique(dst: Path) -> Path:
    """파일명이 이미 존재할 경우 (1), (2) 등을 붙여 고유하게 만듦"""
    if not dst.exists(): return dst
    
    b, e = dst.stem, dst.suffix
    i = 1
    while True:
        c = dst.parent / f"{b}_{i}{e}"
        if not c.exists(): return c
        i += 1

def _natural_sort_key(s):
    """자연 정렬을 위한 키 생성"""
    return [int(s) if s.isdigit() else s.lower() for s in re.split(r'(\d+)', s)]

def list_images(folder: str) -> List[Path]:
    """폴더 내 이미지 파일을 찾아 자연 정렬하여 Path 리스트로 반환"""
    p = Path(folder)
    if not p.is_dir():
        return []
        
    files = [f for f in p.iterdir() 
             if f.is_file() and f.suffix.lower() in IMG_WHITELIST]
    
    # 자연 정렬 적용 (이미지 순서 오류 해결)
    files.sort(key=lambda f: _natural_sort_key(f.name))
    return files

def safe_copy(src: Path, dst: Path):
    """shutil.copy2를 사용한 안전한 복사"""
    shutil.copy2(str(src), str(dst))


# -------------------------------------------------------------
# 2) S1 처리 로직 (작업 사진 - 엑셀 기반 리네임)
# -------------------------------------------------------------

def process_step1(container_dir: Path, s1_folder: Path, s1_excel: Path, clean=True):
    """
    S1 엑셀 파일을 읽어 이미지 파일 목록과 매칭 후 리네임/복사 수행
    """
    
    # 출력 폴더 경로 결정
    out_dir = (s1_folder if OUTPUT_POLICY=="same_as_input" else container_dir / f"{container_dir.name} 작업 사진")
    out_dir.mkdir(parents=True, exist_ok=True)
    
    files = list_images(str(s1_folder))
    
    # 엑셀 데이터 로드: S2 스키마와 동일한 컬럼 인덱스 사용
    COL_SN, COL_ID1, COL_LOT1, COL_TAG1 = 1, 8, 9, 10
    
    try:
        wb = load_workbook(str(s1_excel), data_only=True, read_only=True)
        ws = wb.active
    except Exception:
        return 0, len(files) 

    excel_data = []
    for r in range(2, ws.max_row + 1):
        sn = _nz(ws.cell(row=r, column=COL_SN).value, "NO_SN")
        id_val = _nz(ws.cell(row=r, column=COL_ID1).value, "NO_ID")
        lot = _nz(ws.cell(row=r, column=COL_LOT1).value, "NO_LOT")
        tag = _nz(ws.cell(row=r, column=COL_TAG1).value, "NO_TAG")
        
        if sn != "NO_SN" and lot != "NO_LOT" and tag != "NO_TAG":
            excel_data.append((sn, id_val, lot, tag))

    n = min(len(files), len(excel_data))
    processed_count = 0
    
    for i in range(n):
        sn, id_val, lot, tag = excel_data[i]
        src = s1_folder / files[i].name 
        
        # S1 파일명 구성: SN. ID-LOT (TAG)
        base_name = f"{sn}. {id_val}-{lot} ({tag})"
        
        new_name = sanitize(base_name) + src.suffix
        dst = ensure_unique(out_dir / new_name)
        
        safe_copy(src, dst)
        processed_count += 1
        
    return processed_count, len(files) - processed_count


# -------------------------------------------------------------
# 3) S2 처리 로직 (내품 사진 - 엑셀 기반 리네임)
# -------------------------------------------------------------

def process_step2(container_dir: Path, s2_folder: Path, s2_excel: Path, pair_mode="2", clean=True):
    """
    S2 엑셀 데이터를 읽어 이미지 파일 목록과 매칭 후 리네임/복사 수행
    (이 함수는 main.py의 OCR/백업 로직 후 파일 복사/이동을 담당)
    """
    
    out_dir = (s2_folder if OUTPUT_POLICY=="same_as_input" else container_dir / f"{container_dir.name} 내품사진(LOT NO-TON BAG)")
    out_dir.mkdir(parents=True, exist_ok=True)
    
    files = list_images(str(s2_folder))
    
    # 엑셀 데이터 로드: S2 스키마 기준
    COL_SN, COL_ID1, COL_LOT1, COL_TAG1, COL_TAG2 = 1, 8, 9, 10, 13
    
    try:
        wb = load_workbook(str(s2_excel), data_only=True, read_only=True)
        ws = wb.active
    except Exception:
        return 0, len(files) 

    pair = int(pair_mode) 
    
    # 데이터 행만 추출: SN(1), ID1(8), Lot1(9), Tag1(10), Tag2(13)
    rows_data = []
    for r in range(2, ws.max_row + 1):
        row = (
            _nz(ws.cell(row=r, column=COL_SN).value, "NO_SN"),
            _nz(ws.cell(row=r, column=COL_ID1).value),
            _nz(ws.cell(row=r, column=COL_LOT1).value),
            _nz(ws.cell(row=r, column=COL_TAG1).value),
            _nz(ws.cell(row=r, column=COL_TAG2).value)
        )
        if row[0] != "NO_SN" and row[2] and row[3]: 
             rows_data.append(row)

    idx = 0
    processed_count = 0
    
    for row_tuple in rows_data:
        SN, ID1, Lot1, Tag1, Tag2 = row_tuple

        if idx + pair > len(files): 
            break

        # S2 리네임 로직 (2번 프로그램 기반)
        
        # ------------------ 1 Pack (2샷) 처리 ------------------
        if pair == 2:
            # 샷 1: SN. ID1-LOT1 (TAG1).jpg
            base1 = f"{SN}. {ID1}-{Lot1} ({Tag1})"
            src1 = s2_folder / files[idx].name; idx += 1; processed_count += 1
            new_name1 = sanitize(base1) + src1.suffix
            dst1 = ensure_unique(out_dir / new_name1); safe_copy(src1, dst1)
            
            # 샷 2: SN. ID1-LOT1 (TAG1)-1.jpg
            base2 = f"{SN}. {ID1}-{Lot1} ({Tag1}) - 1"
            src2 = s2_folder / files[idx].name; idx += 1; processed_count += 1
            new_name2 = sanitize(base2) + src2.suffix
            dst2 = ensure_unique(out_dir / new_name2); safe_copy(src2, dst2)
            
        # ------------------ 2 Pack (3샷) 처리 ------------------
        elif pair == 3:
            # 샷 1: Pack1 - SN. ID1-LOT1 (TAG1).jpg
            base1 = f"{SN}. {ID1}-{Lot1} ({Tag1})"
            src1 = s2_folder / files[idx].name; idx += 1; processed_count += 1
            new_name1 = sanitize(base1) + src1.suffix
            dst1 = ensure_unique(out_dir / new_name1); safe_copy(src1, dst1)
            
            # 샷 2: Pack2 - SN. ID1-LOT1 (TAG2).jpg
            base2 = f"{SN}. {ID1}-{Lot1} ({Tag2})"
            src2 = s2_folder / files[idx].name; idx += 1; processed_count += 1
            new_name2 = sanitize(base2) + src2.suffix
            dst2 = ensure_unique(out_dir / new_name2); safe_copy(src2, dst2)
            
            # 샷 3: Stack - SN. ID1-LOT1 (TAG1, TAG2).jpg
            base3 = f"{SN}. {ID1}-{Lot1} ({Tag1}, {Tag2})"
            src3 = s2_folder / files[idx].name; idx += 1; processed_count += 1
            new_name3 = sanitize(base3) + src3.suffix
            dst3 = ensure_unique(out_dir / new_name3); safe_copy(src3, dst3)

    return processed_count, len(files) - processed_count